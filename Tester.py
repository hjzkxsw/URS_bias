'''
Copyright (C) 2026  CIAM Group, Southern University of Science and Technology, Shenzhen, China.

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published
by the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
'''

import os
import torch
from datetime import datetime
from logging import getLogger

from env.UniVRPEnv import UniVRPEnv
from model.Model import Model

from problem.ProblemRep import get_problem_representations
from problem.ProblemSet import ProblemSet

from data.DataFinder import DataFinder
from data.DataReader import get_saved_data

from utils.utils import *
from env.mask.mask_registry import mask_registry

class Tester:
    def __init__(self,
                 env_params,
                 model_params,
                 tester_params):

        # save arguments
        self.env_params = env_params
        self.model_params = model_params
        self.tester_params = tester_params
        
        # result folder, logger
        self.logger = getLogger(name='tester')
        self.result_folder = get_result_folder()
        
        self.test_problem_list = env_params['test_problem_list']
        self.logger.info("{0} problems for testing: {1}".format(len(self.test_problem_list), self.test_problem_list))
        
        self.problem_representation_set = get_problem_representations()
        unknown_problems = [p for p in self.test_problem_list if p not in self.problem_representation_set]
        if unknown_problems:
            raise ValueError(f"Unknown problem names: {unknown_problems}")

        # cuda
        USE_CUDA = self.tester_params['use_cuda']
        if USE_CUDA:
            cuda_device_num = self.tester_params['cuda_device_num']
            torch.cuda.set_device(cuda_device_num)
            device = torch.device('cuda', cuda_device_num)
            torch.set_default_tensor_type('torch.cuda.FloatTensor')
        else:
            device = torch.device('cpu')
            torch.set_default_tensor_type('torch.FloatTensor')
            
        self.device = device

        # Main Components
        self.model = Model(**self.model_params).to(self.device)

        # Restore model from checkpoint
        checkpoint_fullname = tester_params['model_load']
        checkpoint = torch.load(checkpoint_fullname, map_location=device)
        self.model.load_state_dict(checkpoint['model_state_dict'])
        self.logger.info("Model loaded successfully!!!")
        self.logger.info("Model loaded from: {0}".format(checkpoint_fullname))

        total = sum([param.nelement() for param in self.model.parameters()])
        self.logger.info("Number of parameters: %.2fM" % (total / 1e6))

        # utility
        self.time_estimator = TimeEstimator()
        
        # convert problem representations to tensors, which will be used for conditioning the model during training and validation
        self.problem_representation_tensor = {
            name: torch.tensor(rep, dtype=torch.float32, device=self.device)
            for name, rep in self.problem_representation_set.items()
        }
        
        # find the data paths for testing; datasets are loaded lazily per task in run().
        ##############################################################################################################        
        self.data_finder = DataFinder(data_dir=env_params['data_dir'])
        
        self.test_scale_list = env_params['test_scale_list']
        self.test_episodes_list = tester_params['test_episodes']
        self.aug_factor = tester_params['aug_factor']
        self.logger.info("{0} scales for testing: {1}. The corresponding episodes: {2}".format(
            len(self.test_scale_list), self.test_scale_list, self.test_episodes_list))
        self.logger.info("Augmentation enabled: {}, Augmentation factor: {} for symmetric problems, {} for asymmetric problems".format(
            tester_params['augmentation_enable'], self.aug_factor[0], self.aug_factor[1]))

        self.test_data_infos = {}
        for test_scale, test_episodes in zip(self.test_scale_list, self.test_episodes_list):
            self.test_data_infos[test_scale] = {}
            
            for problem_name in self.test_problem_list:
                if test_scale >= 1000 and problem_name in ProblemSet.get(name="large_scale_list"):
                    # For large-scale instances, we replace the oracle from or-tools to pyvrp.
                    self.data_finder.oracle_dict.set(problem_name, test_scale, "pyvrp")
                data_info = self.data_finder.get(problem_name, test_scale)
                if data_info is None:
                    continue # skip missing data file

                self.test_data_infos[test_scale][problem_name] = {
                    **data_info,
                    "episodes": test_episodes,
                }

                self.logger.info(
                    'Found data path for {0:4d} {1:16s} ==> data: {2}, solution: {3}'.format(
                        test_episodes,
                        problem_name.upper()+str(test_scale),
                        data_info["data_file"],
                        data_info["solution_file"],
                    )
                )
            
    # solving synthetic problems, which are randomly generated.
    def run(self):
        self.scale_problem_gaps = {}
        global_start_time = time.time()
        total_count = sum(len(self.test_data_infos.get(test_scale, {})) for test_scale in self.test_scale_list)
        completed_count = 0
        self.logger.info(f"Total test tasks: {total_count}")

        for test_scale in self.test_scale_list:
            if self.test_data_infos[test_scale] == {}:
                continue
            self.scale_problem_gaps[test_scale] = {}
            
            for problem_name in self.test_problem_list:
                if problem_name not in self.test_data_infos[test_scale]:
                    continue

                self.logger.info('========================================================================')
                self.logger.info('========================================================================')

                #########################################################################
                # Test
                #########################################################################
                data_info = self.test_data_infos[test_scale][problem_name]
                test_num_episode = data_info['episodes']
                cur_saved_data, oracle_score = get_saved_data(
                    filename=data_info["data_path"],
                    problem_name=problem_name,
                    total_episodes=test_num_episode,
                    device=self.device,
                    solution_name=data_info["solution_path"],
                )
                self.logger.info('Successfully load {0:4d} {1} instances ==> oracle score: {2:7.4f}'.format(
                    test_num_episode, problem_name.upper()+str(test_scale), oracle_score))
                
                if not self.tester_params['augmentation_enable']:
                    aug_factor = 1
                elif "a" in problem_name:
                    aug_factor = self.aug_factor[1] # 128 for asymmetric problems
                else:
                    aug_factor = self.aug_factor[0] # 8 for symmetric problems

                # set batch size
                batch_size = self._get_test_batch_size(problem_name, test_scale)

                self.logger.info(f"Testing {test_num_episode} {problem_name.upper()}{str(test_scale)} instances with aug_factor: {aug_factor} and batch_size: {batch_size}")

                try:
                    self.test_one_problem(problem_name, cur_saved_data, oracle_score,
                                               test_num_episode, test_scale,aug_factor=aug_factor,
                                               batch_size=batch_size)
                finally:
                    del cur_saved_data
                    self.env = None
                    if self.device.type == "cuda":
                        torch.cuda.empty_cache()

                completed_count += 1
                current_used_time = time.time() - global_start_time
                self.logger.info(f"Completed {completed_count}/{total_count} Tasks. Time used so far: {format_duration(current_used_time)}")
        self._output_test_summary()

    def test_one_problem(self, problem_name, dataset, oracle_score,episodes,problem_size,aug_factor=1,batch_size=None):

        self.model.eval()
        self.model.set_decoder_type(self.model_params['eval_type']) 
        
        # re-instantiate environment for each problem to avoid potential influence of different problem types
        self.env = UniVRPEnv()

        # Get mask function from registry
        mask_fn = mask_registry.build_combined_mask(problem_name)

        self.logger.info(f"Before solution construction, reset time estimator and averageMeter...")
        self.time_estimator.reset()
        score = AverageMeter()
        aug_score = AverageMeter()

        problem_representation = self.problem_representation_tensor[problem_name]
        tested_episodes = 0
        
        if batch_size is None:
            batch_size = episodes

        with torch.inference_mode():
            start_time = time.time()
            while tested_episodes < episodes:
                remaining_episodes = episodes - tested_episodes
                batch_size = min(batch_size, remaining_episodes)
                self.env.load_problems(batch_size=batch_size,
                                       problem_name=problem_name,
                                       problem_size=problem_size,
                                       pomo_size=problem_size,
                                       validation_data=dataset,
                                       aug_factor=aug_factor,
                                       start=tested_episodes,
                                       device=self.device)
                reset_state, _, _ = self.env.reset()

                self.model.decoder.assign(problem_representation)
                self.model.pre_forward(reset_state, problem_name, problem_representation)

                # POMO Rollout
                ###############################################
                state, reward, done = self.env.pre_step()
                while not done:
                    cur_dist = self.env.get_local_feature()
                    selected, _ = self.model(state, cur_dist)
                    # shape: (batch, pomo)
                    state, reward, done = self.env.step(selected,mask_fn=mask_fn)


                # Return
                ###############################################
                aug_reward = reward.reshape(aug_factor, batch_size, self.env.pomo_size)
                # shape: (augmentation, batch, pomo)

                max_pomo_reward, _ = aug_reward.max(dim=2)  # get best results from pomo
                # shape: (augmentation, batch)
                if problem_name in ["op","aop"]:
                    avg_score = max_pomo_reward[0, :].float().mean().item()
                else:
                    avg_score = -max_pomo_reward[0, :].float().mean().item()  # negative sign to make positive value

                max_aug_pomo_reward, _ = max_pomo_reward.max(dim=0)  # get best results from augmentation
                # shape: (batch,)
                if problem_name in ["op",'aop']:
                    avg_aug_score = max_aug_pomo_reward.float().mean().item()
                else:
                    avg_aug_score = -max_aug_pomo_reward.float().mean().item()  # negative sign to make positive value
                    # shape: (1,)
                
                assert max_aug_pomo_reward.shape[0] == batch_size, \
                    f"Expected reward shape: ({batch_size},), but got {max_aug_pomo_reward.shape}"

                score.update(avg_score, batch_size)
                aug_score.update(avg_aug_score, batch_size)
                
                tested_episodes += batch_size

                ############################
                # Logs
                ############################
                elapsed_time_str, remain_time_str = self.time_estimator.get_est_string(tested_episodes, episodes)
                self.logger.info("episode {:4d}/{:4d}, Elapsed[{}], Remain[{}], score:{:.3f}, aug_score:{:.3f}".format(
                        tested_episodes, episodes, elapsed_time_str, remain_time_str, avg_score, avg_aug_score))

                all_done = (tested_episodes == episodes)
                if all_done:
                    end_time = time.time()
                    during_time = end_time - start_time
                    if problem_name in ["op",'aop']:
                        gap = ((oracle_score - score.avg) * 100 / oracle_score)
                        aug_gap = ((oracle_score - aug_score.avg) * 100 / oracle_score)
                    else:
                        gap = ((score.avg - oracle_score) * 100 / oracle_score)
                        aug_gap = ((aug_score.avg - oracle_score) * 100 / oracle_score)
                    self.logger.info(f"************* Test Done: {problem_name.upper()}{str(problem_size)} ************* ")
                    self.logger.info("problem size: {0}, pomo size: {1}, aug factor: {2}, oracle score: {3:.4f} ".format(
                            problem_size, self.env.pomo_size, aug_factor, oracle_score))
                    self.logger.info("NO-AUG SCORE:{0:.4f}, GAP:{1:.3f}%".format(
                        score.avg, gap))
                    self.logger.info("AUGMENTATION SCORE:{0:.4f}, GAP:{1:.3f}%".format(
                        aug_score.avg, aug_gap))

                    self.logger.info("Total time: {:.2f} sec, {:.2f} min".format(during_time, during_time / 60))
                    self.logger.info("Avg time per episode: {:.2f} sec".format(during_time / episodes))

                    # Update best gap
                    self.scale_problem_gaps[problem_size][problem_name] = {
                        'no_aug_gap': gap,
                        'aug_gap': aug_gap,
                        "time": during_time,
                        "avg_time": during_time / episodes,
                    }
    
    def _get_test_batch_size(self, problem_name, test_scale):
        if test_scale >= 1000:
            large_scale_to_batch_size = dict(zip([1000, 2000, 3000, 4000, 5000], self.tester_params["test_batch_size_large"]))

            if test_scale not in large_scale_to_batch_size:
                raise ValueError(f"No batch size configured for large scale: {test_scale}")

            return large_scale_to_batch_size[test_scale]

        small_batch_sizes = self.tester_params["test_batch_size_small"]

        if "a" in problem_name:
            return small_batch_sizes[3] if "md" in problem_name else small_batch_sizes[2]

        return small_batch_sizes[1] if "md" in problem_name else small_batch_sizes[0]
    
    def _collect_test_summary_rows(self):
        mode = "Aug" if self.tester_params['augmentation_enable'] else "No Aug"
        rows = []

        for test_scale, test_episodes in zip(self.test_scale_list, self.test_episodes_list):
            scale_results = self.scale_problem_gaps.get(test_scale, {})

            for problem_name in self.test_problem_list:
                if problem_name not in scale_results:
                    continue

                result = scale_results[problem_name]

                rows.append({
                    "scale": test_scale,
                    "problem": problem_name,
                    "episodes": test_episodes,
                    "mode": mode,
                    "gap": f"{to_float(result['aug_gap']):.3f}%",
                    "total_time": format_duration(to_float(result["time"]),demical_places=1),
                    "avg_time": format_duration(to_float(result["avg_time"]),demical_places=2),
                })
        return rows

    def _format_test_summary_blocks(self, rows, problems_per_row):
        if problems_per_row <= 0:
            raise ValueError("summary_problems_per_row must be greater than 0")

        log_lines = []
        block_groups = []
        rows_by_scale = {}
        global_block_width = 0

        for row in rows:
            rows_by_scale.setdefault(row["scale"], []).append(row)
            half_width = max(
                len("Gap"),
                len("Time"),
                len(row["gap"]),
                len(row["total_time"]),
            )
            global_block_width = max(
                global_block_width,
                len(row["problem"].upper()),
                half_width * 2,
            )

        if global_block_width % 2:
            global_block_width += 1

        for test_scale in self.test_scale_list:
            scale_rows = rows_by_scale.get(test_scale, [])
            if not scale_rows:
                continue

            log_lines.append(f"### Scale {test_scale}")
            scale_groups = []

            for start in range(0, len(scale_rows), problems_per_row):
                row_group = scale_rows[start:start + problems_per_row]
                blocks = [self._build_summary_block(row, global_block_width) for row in row_group]
                scale_groups.append(blocks)
                group_lines = self._format_summary_block_group(blocks)
                log_lines.extend(group_lines)
                log_lines.append(self._build_summary_separator(group_lines))

            block_groups.append({
                "scale": test_scale,
                "groups": scale_groups,
            })

        return log_lines, block_groups

    def _build_summary_block(self, row, block_width):
        problem = row["problem"].upper()
        gap = row["gap"]
        time_display = row["total_time"]
        gap_width = block_width // 2
        time_width = block_width // 2

        return {
            "problem": problem,
            "gap": gap,
            "time": time_display,
            "gap_width": gap_width,
            "time_width": time_width,
            "block_width": block_width,
            "problem_line": problem.center(block_width),
            "label_line": f"{'Gap':>{gap_width}}{'Time':>{time_width}}",
            "value_line": f"{gap:>{gap_width}}{time_display:>{time_width}}",
        }

    def _format_summary_block_group(self, blocks):
        return [
            self._join_summary_block_line(blocks, "problem_line"),
            self._join_summary_block_line(blocks, "label_line"),
            self._join_summary_block_line(blocks, "value_line"),
        ]

    def _join_summary_block_line(self, blocks, key):
        return "| " + " | ".join(block[key] for block in blocks) + " |"

    def _build_summary_separator(self, group_lines):
        width = max(len(line) for line in group_lines)
        dash_count = max(2, (width + 1) // 2)
        return " ".join(["—"] * dash_count)

    def _save_test_summary_excel(self, rows, block_groups):
        import pandas as pd

        os.makedirs(self.result_folder, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        excel_path = os.path.join(self.result_folder, f"test_summary_{timestamp}.xlsx")

        summary_columns = [
            "scale",
            "problem",
            "episodes",
            "mode",
            "gap",
            "total_time",
            "avg_time",
        ]
        summary_df = pd.DataFrame(rows, columns=summary_columns)

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="summary", index=False)
            self._write_block_layout_sheet(writer.book, block_groups)

        return excel_path

    def _write_block_layout_sheet(self, workbook, block_groups):
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        sheet = workbook.create_sheet("block_layout")
        center_alignment = Alignment(horizontal="center", vertical="center")
        header_font = Font(bold=True)
        current_row = 1

        for scale_block in block_groups:
            sheet.cell(row=current_row, column=1, value=f"Scale {scale_block['scale']}")
            sheet.cell(row=current_row, column=1).font = header_font
            current_row += 1

            for group in scale_block["groups"]:
                current_col = 1

                for block in group:
                    sheet.merge_cells(
                        start_row=current_row,
                        start_column=current_col,
                        end_row=current_row,
                        end_column=current_col + 1,
                    )
                    sheet.cell(row=current_row, column=current_col, value=block["problem"])
                    sheet.cell(row=current_row + 1, column=current_col, value="Gap")
                    sheet.cell(row=current_row + 1, column=current_col + 1, value="Time")
                    sheet.cell(row=current_row + 2, column=current_col, value=block["gap"])
                    sheet.cell(row=current_row + 2, column=current_col + 1, value=block["time"])

                    for row_index in range(current_row, current_row + 3):
                        for col_index in range(current_col, current_col + 2):
                            sheet.cell(row=row_index, column=col_index).alignment = center_alignment

                    sheet.cell(row=current_row, column=current_col).font = header_font
                    current_col += 2

                current_row += 3

        for col_index in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_index)].width = 14

    def _output_test_summary(self):
        rows = self._collect_test_summary_rows()
        mode = "Augmentation" if self.tester_params['augmentation_enable'] else "No Augmentation"
        problems_per_row = self.tester_params.get("summary_problems_per_row", 4)
        log_lines, block_groups = self._format_test_summary_blocks(rows, problems_per_row)

        self.logger.info('=============================== Summary of experimental results ===================================')
        self.logger.info(f"Mode: {mode}")

        for line in log_lines:
            self.logger.info(line)

        excel_path = self._save_test_summary_excel(rows, block_groups)
        self.logger.info(f"Summary excel is saved to: {excel_path}")
         
